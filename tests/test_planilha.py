import copy
import os

import openpyxl

from bolao.planilha import (
    ler_jogos, ler_palpites, ler_resultados, escrever_resultado,
    escrever_palpites, reordenar_classificacao, adicionar_jogos,
    estender_somas_total_pontos,
)
from bolao.pontuacao import montar_ranking
from bolao.modelo import PARTICIPANTES, BONUS, JOGOS_IGNORADOS, ULTIMA_LINHA

# Snapshot CONGELADO da planilha (estado "até 24/jun"). Estes testes validam o
# motor de pontuação contra dados conhecidos — por isso NÃO podem ler a planilha
# mestre viva (`TABELA_APOSTAS_-_COPA_2026.xlsx`), que o robô atualiza todo dia e
# faria o gold test quebrar a cada rodada. Fixture extraído via git (commit eaae127);
# regenerar de propósito só se as regras/engine mudarem.
ARQUIVO = os.path.join(os.path.dirname(__file__), "fixtures", "planilha_oficial_ate_24jun.xlsx")

# Totais oficiais do snapshot (aba "Total Pontos", resultados preenchidos até 24/jun).
ESPERADO = {
    "Mané": 110, "AH": 107, "Romanelli": 106, "Su": 100, "Rodrigo": 99,
    "Beda": 95, "Mauro": 95, "Pedro": 90, "Caio": 89, "Biral": 87,
    "Paulo": 85, "Camps": 74, "Kim": 58,
}


def _carregar():
    return openpyxl.load_workbook(ARQUIVO, data_only=True)


def test_reproduz_totais_oficiais():
    wb = _carregar()
    resultados = ler_resultados(wb)
    palpites = {disp: ler_palpites(wb, aba) for disp, aba in PARTICIPANTES}
    ranking = dict(montar_ranking(palpites, resultados, BONUS, ignorar=JOGOS_IGNORADOS))
    assert ranking == ESPERADO


def test_jogos_ignorados_nao_afetam_ranking():
    # Mesmo se os jogos 1-4 tiverem placar, eles NÃO podem contar (decisão do time).
    wb = openpyxl.load_workbook(ARQUIVO)
    for num in (1, 2, 3, 4):
        escrever_resultado(wb, num, 5, 0)  # placar arbitrário
    resultados = ler_resultados(wb)
    palpites = {disp: ler_palpites(wb, aba) for disp, aba in PARTICIPANTES}
    ranking = dict(montar_ranking(palpites, resultados, BONUS, ignorar=JOGOS_IGNORADOS))
    assert ranking == ESPERADO  # inalterado


def test_ler_jogos_tem_72_e_brasil_escocia():
    wb = _carregar()
    jogos = ler_jogos(wb)
    assert len(jogos) == 72
    # jogo 51: Escócia x Brasil em 24/jun
    assert jogos[51].time1 == "Escócia"
    assert jogos[51].time2 == "Brasil"
    assert jogos[51].data == "2026-06-24"


def test_escrever_resultado_nao_sobrescreve():
    wb = openpyxl.load_workbook(ARQUIVO)  # com fórmulas
    # jogo 55 (Equador x Alemanha, 25/jun) está vazio -> deve escrever
    assert escrever_resultado(wb, 55, 1, 2) is True
    # tentar de novo NÃO sobrescreve
    assert escrever_resultado(wb, 55, 9, 9) is False
    # jogo 51 já tem placar -> não escreve
    assert escrever_resultado(wb, 51, 0, 0) is False
    res = ler_resultados(wb)
    assert res[55] == (1, 2)
    assert res[51] == (0, 3)  # Escócia 0 x 3 Brasil permanece


def test_reordenar_classificacao_escreve_ordem_e_formulas():
    wb = openpyxl.load_workbook(ARQUIVO)
    ranking = [("Mané", 110), ("AH", 107), ("Kim", 58)] + \
              [(n, 0) for n, _ in PARTICIPANTES if n not in ("Mané", "AH", "Kim")]
    reordenar_classificacao(wb, ranking)
    ws = wb["Total Pontos"]
    assert ws.cell(12, 4).value == "Mané"
    assert ws.cell(13, 4).value == "AH"
    # Kim tem bônus +10 na fórmula; a faixa cobre o torneio todo (até o cap)
    assert ws.cell(14, 4).value == "Kim"
    assert ws.cell(14, 5).value == f"=SUM('Kim'!K5:K{ULTIMA_LINHA})+10"
    # Mané não tem bônus
    assert ws.cell(12, 5).value == f"=SUM('Mané'!K5:K{ULTIMA_LINHA})"


def test_escrever_palpites_grava_gols_nas_colunas_D_e_F():
    # Merge do form -> aba do participante. Jogo do mata-mata (ex.: nº 73).
    wb = openpyxl.Workbook()
    wb.active.title = "Caio"
    escrever_palpites(wb, "Caio", {73: (0, 2), 74: (1, 1)})
    ws = wb["Caio"]
    assert (ws.cell(77, 4).value, ws.cell(77, 6).value) == (0, 2)  # jogo 73 -> linha 77 (D/F)
    assert (ws.cell(78, 4).value, ws.cell(78, 6).value) == (1, 1)  # jogo 74 -> linha 78


def test_escrever_palpites_sobrescreve_palpite_anterior():
    # Caio re-mergeia porque o participante mudou a aposta antes do prazo.
    wb = openpyxl.Workbook()
    wb.active.title = "Su"
    escrever_palpites(wb, "Su", {73: (3, 0)})
    escrever_palpites(wb, "Su", {73: (1, 2)})
    ws = wb["Su"]
    assert (ws.cell(77, 4).value, ws.cell(77, 6).value) == (1, 2)


# --- Fase 2: estender a planilha com os jogos do mata-mata ---

def test_adicionar_jogos_insere_em_resultados_e_nas_abas():
    wb = openpyxl.Workbook()
    wb.active.title = "Resultados"
    wb.create_sheet("Beda")
    wb.create_sheet("Su")
    adicionar_jogos(wb, [
        {"numero": 73, "rodada": "32-avos", "time1": "Brasil",
         "time2": "Croácia", "data": "2026-06-28", "hora": "16:00"},
    ])
    # o jogo entra na aba Resultados, SEM placar oficial
    jogos = ler_jogos(wb)
    assert 73 in jogos
    assert (jogos[73].time1, jogos[73].time2) == ("Brasil", "Croácia")
    assert jogos[73].data == "2026-06-28"
    assert ler_resultados(wb)[73] == (None, None)
    # cada aba de participante recebe a fórmula K (linha 77) e palpite vazio
    for aba in ("Beda", "Su"):
        ws = wb[aba]
        assert ws.cell(77, 11).value.startswith("=IF(AND(ISNUMBER(D77)")
        assert (ws.cell(77, 4).value, ws.cell(77, 6).value) == (None, None)


def test_adicionar_jogos_usa_so_as_abas_presentes():
    # não deve falhar se nem todas as 13 abas existem
    wb = openpyxl.Workbook()
    wb.active.title = "Resultados"
    wb.create_sheet("Beda")
    adicionar_jogos(wb, [{"numero": 73, "rodada": "R32", "time1": "A",
                          "time2": "B", "data": "2026-06-28", "hora": "16:00"}])
    assert wb["Beda"].cell(77, 11).value is not None


def test_estender_somas_total_pontos_ate_o_cap():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Total Pontos"
    ws.cell(6, 4).value = "=SUM(Beda!K5:K76)"          # tabela horizontal (linha 6)
    ws.cell(21, 5).value = "=SUM('Caio'!K5:K76)+10"     # com bônus
    n = estender_somas_total_pontos(wb)
    assert n == 2
    assert ws.cell(6, 4).value == f"=SUM(Beda!K5:K{ULTIMA_LINHA})"
    assert ws.cell(21, 5).value == f"=SUM('Caio'!K5:K{ULTIMA_LINHA})+10"
    assert estender_somas_total_pontos(wb) == 0  # idempotente
