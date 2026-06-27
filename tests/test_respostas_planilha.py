"""Leitura dos arquivos .xlsx preenchidos por cada participante (mata-mata).

Diferente do Google Forms (uma planilha com todas as respostas), aqui cada um
devolve o PRÓPRIO arquivo, no formato do molde (`scripts/gerar_molde_mata_mata.py`):
col A=nº do jogo (1..N), D=gols1, F=gols2. O Caio renomeia cada retorno com o
nome no fim (ex.: "APOSTAS_MATA-MATA_COPA_2026 Kim.xlsx").

Regras:
- jogo só conta se os DOIS gols estiverem preenchidos (em branco != 0x0);
- aceita gols como número ou texto;
- o dono do arquivo vem do fim do nome do arquivo (à prova de acento/maiúscula);
- no merge, o jogo local i (1..N) vira o jogo (primeiro_jogo + i - 1) na mestre.
"""
import openpyxl
import pytest

from bolao.planilha import ler_palpites
from bolao.respostas_planilha import (
    ler_palpites_arquivo, participante_do_arquivo, merge_arquivos,
)


def _wb_arquivo(jogos):
    """Monta um arquivo no formato do molde. `jogos`: lista de itens
    (num, gols1, gols2) preenchidos, ou (num,) para jogo em branco."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Apostas"
    for i, item in enumerate(jogos):
        r = 5 + i
        ws.cell(r, 1).value = item[0]   # A = nº do jogo
        ws.cell(r, 5).value = "X"       # E = separador
        if len(item) == 3:
            ws.cell(r, 4).value = item[1]   # D = gols1
            ws.cell(r, 6).value = item[2]   # F = gols2
    return wb


def test_le_palpites_de_arquivo_preenchido():
    wb = _wb_arquivo([(1, 2, 1), (2, 0, 0), (3, 3, 2)])
    assert ler_palpites_arquivo(wb) == {1: (2, 1), 2: (0, 0), 3: (3, 2)}


def test_jogo_em_branco_ou_pela_metade_nao_conta():
    # jogo 2 totalmente vazio; jogo 3 só com um gol -> nenhum dos dois conta
    wb = _wb_arquivo([(1, 1, 0), (2,), (3, 2, None)])
    assert ler_palpites_arquivo(wb) == {1: (1, 0)}


def test_aceita_gols_texto_e_numero():
    wb = _wb_arquivo([(1, "2", 1), (2, 0.0, "3")])
    assert ler_palpites_arquivo(wb) == {1: (2, 1), 2: (0, 3)}


@pytest.mark.parametrize("arquivo, esperado", [
    ("APOSTAS_MATA-MATA_COPA_2026 Kim.xlsx", "Kim"),
    ("APOSTAS_MATA-MATA_COPA_2026 Camps.xlsx", "Camps "),   # aba tem espaço no fim
    ("APOSTAS_MATA-MATA_COPA_2026 Mane.xlsx", "Mané"),      # arquivo sem acento
    ("APOSTAS_MATA-MATA_COPA_2026 mané.xlsx", "Mané"),      # minúsculo + acento
    (r"C:\pasta\APOSTAS_MATA-MATA_COPA_2026 Su.xlsx", "Su"),  # caminho completo
    ("qualquer_coisa_sem_nome.xlsx", None),
])
def test_participante_do_arquivo(arquivo, esperado):
    assert participante_do_arquivo(arquivo) == esperado


def test_merge_arquivos_grava_na_aba_certa_com_offset(tmp_path):
    mestre = openpyxl.Workbook()
    mestre.remove(mestre.active)
    for aba in ("Kim", "Su"):
        mestre.create_sheet(aba)

    f_kim = tmp_path / "APOSTAS_MATA-MATA_COPA_2026 Kim.xlsx"
    _wb_arquivo([(1, 2, 1), (2, 0, 0)]).save(f_kim)
    f_su = tmp_path / "APOSTAS_MATA-MATA_COPA_2026 Su.xlsx"
    _wb_arquivo([(1, 3, 3)]).save(f_su)
    f_estranho = tmp_path / "arquivo_sem_dono.xlsx"
    _wb_arquivo([(1, 9, 9)]).save(f_estranho)

    gravados, sem_dono = merge_arquivos(
        mestre, [str(f_kim), str(f_su), str(f_estranho)], primeiro_jogo=73,
    )

    assert gravados == {"Kim": 2, "Su": 1}
    assert sem_dono == [str(f_estranho)]
    # local 1 -> jogo 73, local 2 -> jogo 74
    assert ler_palpites(mestre, "Kim")[73] == (2, 1)
    assert ler_palpites(mestre, "Kim")[74] == (0, 0)
    assert ler_palpites(mestre, "Su")[73] == (3, 3)
