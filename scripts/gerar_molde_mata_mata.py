"""Gera o MOLDE em branco do mata-mata a partir da planilha mestre.

Ideia (sugestão do Caio): em vez de recriar a formatação do zero, a gente
DUPLICA uma aba de participante que já existe (título verde, "Nome/Celular/
e-mail", cabeçalho verde, cores/fontes/larguras) e só edita o que muda de uma
fase pra outra — os jogos. Assim a formatação sai idêntica.

O molde é o arquivo que o Caio manda no grupo. Cada participante preenche os
gols (cols D/F) e devolve; o Caio renomeia cada retorno com o nome no fim
(ex.: "APOSTAS_MATA-MATA_COPA_2026 Kim.xlsx") e o merge na mestre é feito por
`bolao/respostas_planilha.py`.

Decisões do molde:
- 1 aba só, N linhas de jogos numeradas 1..N (R32=16, R16=8, QF=4, SF=2, etc.);
- times/data/hora EM BRANCO (preenchidos só quando os confrontos saírem);
- gols (D/F) EM BRANCO — é o que cada um preenche — com o "X" no meio;
- coluna GRUPO em branco (mata-mata não tem grupo);
- SEM a coluna K (pontos): a fórmula dela aponta pra aba "Resultados", que não
  vai junto no molde; ficaria #REF pro participante. A pontuação acontece na
  mestre, no merge.

Uso:
    python scripts/gerar_molde_mata_mata.py            # 16 jogos, nome padrão
    python scripts/gerar_molde_mata_mata.py --jogos 8 --titulo "... OITAVAS ..."
"""
import argparse
from copy import copy
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill

from bolao.modelo import PARTICIPANTES, PRIMEIRA_LINHA
from bolao.planilha import COL, COL_SEP, COL_DATA2, COL_PONTOS

MESTRE_PADRAO = "TABELA_APOSTAS_-_COPA_2026.xlsx"
SAIDA_PADRAO = "APOSTAS_MATA-MATA_COPA_2026.xlsx"
TITULO_PADRAO = "COPA 2026 - APOSTAS MATA-MATA (1ª RODADA - 16 JOGOS)"

# Nomes em PT pra exibir DIA/dia-da-semana no molde como strings (à prova do
# locale do Excel: o molde é só display, o merge lê só os gols).
MESES_PT = ["janeiro", "fevereiro", "março", "abril", "maio", "junho", "julho",
            "agosto", "setembro", "outubro", "novembro", "dezembro"]
DIAS_PT = ["segunda-feira", "terça-feira", "quarta-feira", "quinta-feira",
           "sexta-feira", "sábado", "domingo"]
# Time que recebe o destaque verde-e-amarelo (a aba-base já traz esse destaque
# herdado de onde o Brasil jogou na fase de grupos; aqui ele é realocado pra
# célula do Brasil na rodada atual, onde quer que ele esteja).
DESTAQUE_TIME = "Brasil"
AMARELO = "FFFFFF00"   # fill do destaque (amarelo)
VERDE = "FF00B050"     # fonte do destaque (verde)


def _aba_base(wb):
    """Primeira aba de participante existente, pra servir de molde."""
    for _disp, aba in PARTICIPANTES:
        if aba in wb.sheetnames:
            return aba
    raise ValueError("Nenhuma aba de participante encontrada na mestre.")


def _escrever_data(ws, r, data_iso):
    """DIA (col H) = '28 de junho' e dia-da-semana (col I) = 'domingo', como
    strings em PT. Sem `data_iso`, deixa as duas em branco."""
    cel_dia = ws.cell(r, COL["data"])
    cel_sem = ws.cell(r, COL_DATA2)
    if data_iso:
        d = datetime.strptime(data_iso, "%Y-%m-%d").date()
        cel_dia.value = f"{d.day} de {MESES_PT[d.month - 1]}"
        cel_sem.value = DIAS_PT[d.weekday()]
        cel_dia.number_format = "General"
        cel_sem.number_format = "General"
    else:
        cel_dia.value = None
        cel_sem.value = None


def _eh_amarela(cel):
    """True se a célula tem o fill amarelo do destaque (à prova de cor de tema,
    cujo .rgb levanta exceção)."""
    try:
        return cel.fill.patternType == "solid" and cel.fill.fgColor.rgb == AMARELO
    except (TypeError, ValueError, AttributeError):
        return False


def _destacar_time(ws, primeira, n_jogos, confrontos, time=DESTAQUE_TIME):
    """Realoca o destaque verde-e-amarelo pra célula do `time` na rodada atual.

    A aba-base traz o destaque herdado (de onde o Brasil jogou nos grupos) numa
    linha fixa — que agora pode cair em outro time. Aqui: captura o estilo do
    destaque e o estilo default, zera o destaque herdado nas células erradas e
    aplica na célula do `time`. Se o `time` não estiver na rodada, só limpa."""
    cols = (COL["time1"], COL["time2"])
    fill_dest = font_dest = fill_def = font_def = None
    herdadas = []
    for i in range(n_jogos):
        r = primeira + i
        for col in cols:
            cel = ws.cell(r, col)
            if _eh_amarela(cel):
                herdadas.append((r, col))
                if fill_dest is None:
                    fill_dest, font_dest = copy(cel.fill), copy(cel.font)
            elif fill_def is None:
                fill_def, font_def = copy(cel.fill), copy(cel.font)

    alvo = None
    for i in range(n_jogos):
        c = confrontos[i]
        if c.get("time1") == time:
            alvo = (primeira + i, COL["time1"])
        elif c.get("time2") == time:
            alvo = (primeira + i, COL["time2"])

    # fallback: se a aba-base não trouxer mais o amarelo, constrói o estilo
    if alvo is not None and fill_dest is None:
        base = ws.cell(*alvo).font
        fill_dest = PatternFill("solid", fgColor=AMARELO)
        font_dest = Font(name=base.name, size=base.size, bold=True, color=VERDE)

    # zera o destaque herdado (menos na própria célula-alvo)
    if fill_def is not None:
        for rc in herdadas:
            if rc != alvo:
                ws.cell(*rc).fill = copy(fill_def)
                ws.cell(*rc).font = copy(font_def)

    # aplica o destaque na célula do time-alvo
    if alvo is not None and fill_dest is not None:
        ws.cell(*alvo).fill = copy(fill_dest)
        ws.cell(*alvo).font = copy(font_dest)


def gerar_molde(mestre_path=MESTRE_PADRAO, saida=SAIDA_PADRAO,
                n_jogos=16, titulo=TITULO_PADRAO, confrontos=None):
    """Cria o molde e salva em `saida`. Retorna o caminho salvo.

    Sem `confrontos`: molde EM BRANCO (times/data/hora vazios).
    Com `confrontos` (lista de dicts {time1, time2, data, hora, rodada?}, na
    ordem do chaveamento): preenche os times/data/hora — pronto pra mandar.
    Os gols (D/F) ficam SEMPRE em branco (é o que cada um preenche)."""
    if confrontos:
        n_jogos = len(confrontos)

    wb = openpyxl.load_workbook(mestre_path)
    base = _aba_base(wb)

    # 1) fica só com a aba molde
    for nome in list(wb.sheetnames):
        if nome != base:
            del wb[nome]
    ws = wb[base]
    ws.title = "Apostas"

    # 2) título e limpeza do cabeçalho (nome/celular/e-mail em branco)
    ws.cell(1, 1).value = titulo
    for coord in ("B2", "B3", "G3"):  # valores de Nome / Celular / e-mail
        ws[coord].value = None

    # 3) reescreve as N linhas de jogos (gols sempre em branco)
    primeira = PRIMEIRA_LINHA
    ultima = primeira + n_jogos - 1
    for i in range(n_jogos):
        r = primeira + i
        c = confrontos[i] if confrontos else {}
        ws.cell(r, COL["num"]).value = i + 1
        ws.cell(r, COL["grupo"]).value = c.get("rodada")
        ws.cell(r, COL["time1"]).value = c.get("time1")
        ws.cell(r, COL["gols1"]).value = None
        ws.cell(r, COL_SEP).value = "X"
        ws.cell(r, COL["gols2"]).value = None
        ws.cell(r, COL["time2"]).value = c.get("time2")
        _escrever_data(ws, r, c.get("data"))
        ws.cell(r, COL["hora"]).value = c.get("hora")
        ws.cell(r, COL_PONTOS).value = None  # sem coluna de pontos no molde

    # 3b) realoca o destaque verde-e-amarelo pra célula do Brasil desta rodada
    if confrontos:
        _destacar_time(ws, primeira, n_jogos, confrontos)

    # 4) apaga linhas de jogos sobrando abaixo da última
    if ws.max_row > ultima:
        ws.delete_rows(ultima + 1, ws.max_row - ultima)

    # 5) limpa nomes definidos (apontariam p/ abas apagadas e travariam o Excel)
    try:
        ws_defined = ws.defined_names
        for nome in list(ws_defined):
            del ws_defined[nome]
    except Exception:
        pass
    try:
        for nome in list(wb.defined_names):
            del wb.defined_names[nome]
    except Exception:
        pass

    wb.save(saida)
    return saida


def main():
    p = argparse.ArgumentParser(description="Gera o molde em branco do mata-mata.")
    p.add_argument("--mestre", default=MESTRE_PADRAO)
    p.add_argument("--saida", default=SAIDA_PADRAO)
    p.add_argument("--jogos", type=int, default=16)
    p.add_argument("--titulo", default=TITULO_PADRAO)
    a = p.parse_args()
    caminho = gerar_molde(a.mestre, a.saida, a.jogos, a.titulo)
    print(f"Molde gerado: {caminho}")


if __name__ == "__main__":
    main()
