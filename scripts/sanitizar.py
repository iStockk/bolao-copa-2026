"""Uso único: remove contato pessoal (nome completo, telefone, e-mail) da
cópia da planilha que vai para o repositório público. A planilha local do
usuário permanece intacta — isto roda só sobre a cópia versionada.

Remove tanto o VALOR quanto o HYPERLINK das células de contato (o e-mail/telefone
costumam estar embutidos num hyperlink mailto:/tel: que sobrevive a limpar só o valor),
e faz uma varredura de segurança em todas as abas removendo qualquer hyperlink
mailto:/tel: e qualquer valor de texto que contenha '@' (e-mail).
"""
import re
import sys
import openpyxl

ARQUIVO = "TABELA_APOSTAS_-_COPA_2026.xlsx"
# sequência longa de dígitos com separadores comuns de telefone
_RE_TEL = re.compile(r"\d[\d\s().+-]{7,}\d")
ABAS_PARTICIPANTES = [
    "Beda", "Pedro", "Mauro", "Rodrigo", "Paulo", "Su", "Biral",
    "Mané", "Caio", "Camps ", "Romanelli", "AH", "Kim",
]
CELULAS_CONTATO = ("B2", "B3", "G3")


def _limpar_celula(cel):
    """Zera valor e hyperlink de uma célula. Retorna True se algo foi limpo."""
    mudou = False
    if cel.value is not None:
        cel.value = None
        mudou = True
    if cel.hyperlink is not None:
        cel.hyperlink = None
        mudou = True
    return mudou


def main(arquivo=ARQUIVO):
    wb = openpyxl.load_workbook(arquivo)
    limpas = 0

    # 1) Células de contato conhecidas nas abas de participantes.
    for aba in ABAS_PARTICIPANTES:
        if aba not in wb.sheetnames:
            print(f"AVISO: aba '{aba}' não encontrada", file=sys.stderr)
            continue
        ws = wb[aba]
        for cel in CELULAS_CONTATO:
            if _limpar_celula(ws[cel]):
                limpas += 1

    # 2) Varredura de segurança em TODAS as abas: remove hyperlinks mailto:/tel:
    #    e valores de texto que pareçam e-mail.
    for ws in wb.worksheets:
        for linha in ws.iter_rows():
            for cel in linha:
                hl = cel.hyperlink
                if hl is not None:
                    alvo = getattr(hl, "target", "") or ""
                    if alvo.lower().startswith(("mailto:", "tel:")):
                        cel.hyperlink = None
                        limpas += 1
                v = cel.value
                eh_email = isinstance(v, str) and "@" in v and "." in v.split("@")[-1]
                eh_tel_txt = (isinstance(v, str) and _RE_TEL.search(v)
                              and sum(c.isdigit() for c in v) >= 9)
                eh_tel_num = (isinstance(v, int) and not isinstance(v, bool)
                              and v >= 10 ** 8)
                if eh_email or eh_tel_txt or eh_tel_num:
                    print(f"  - removido {ws.title}!{cel.coordinate}", file=sys.stderr)
                    cel.value = None
                    limpas += 1

    wb.save(arquivo)
    print(f"OK: {limpas} item(ns) de contato removido(s) em {arquivo}")


if __name__ == "__main__":
    main()
