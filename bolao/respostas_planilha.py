"""Leitura dos arquivos .xlsx preenchidos por cada participante (mata-mata).

Alternativa ao Google Forms: cada um devolve o PRÓPRIO arquivo, no formato do
molde (`scripts/gerar_molde_mata_mata.py`) — col A=nº do jogo (1..N), D=gols1,
F=gols2. O Caio renomeia cada retorno com o nome no fim (ex.:
"APOSTAS_MATA-MATA_COPA_2026 Kim.xlsx") e joga todos numa pasta.

`merge_arquivos` lê a pasta e grava os palpites nas abas da planilha mestre.
A escrita em si usa `planilha.escrever_palpites`; a pontuação (inclusive a regra
dos pênaltis) continua no motor de sempre.
"""
import os

import openpyxl

from bolao.modelo import PARTICIPANTES, PRIMEIRA_LINHA, ULTIMA_LINHA
from bolao.nomes import normalizar
from bolao.planilha import COL, escrever_palpites


def _para_int(v):
    """Converte gols (texto '3', float 3.0 ou int 3) para int; vazio -> None."""
    if v is None or v == "":
        return None
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return None


def _aba_jogos(wb):
    """A aba com os jogos é a que tem um número na coluna A da 1ª linha de jogo."""
    for ws in wb.worksheets:
        v = ws.cell(PRIMEIRA_LINHA, COL["num"]).value
        if isinstance(v, (int, float)):
            return ws
    raise ValueError("Não achei a aba de jogos (coluna A da linha 5 sem número).")


def ler_palpites_arquivo(wb):
    """Lê um arquivo preenchido -> {num_jogo: (gols1, gols2)}.

    O nº do jogo vem da coluna A. Um jogo só entra se os DOIS gols estiverem
    preenchidos (em branco != 0x0; meio-preenchido é ignorado)."""
    ws = _aba_jogos(wb)
    palpites = {}
    for r in range(PRIMEIRA_LINHA, ULTIMA_LINHA + 1):
        num = ws.cell(r, COL["num"]).value
        if not isinstance(num, (int, float)):
            continue
        g1 = _para_int(ws.cell(r, COL["gols1"]).value)
        g2 = _para_int(ws.cell(r, COL["gols2"]).value)
        if g1 is not None and g2 is not None:
            palpites[int(num)] = (g1, g2)
    return palpites


def participante_do_arquivo(caminho):
    """Descobre o dono pelo fim do nome do arquivo -> nome da ABA na mestre.

    À prova de acento e maiúscula/minúscula. Retorna None se não reconhecer
    (aí o merge avisa, e o Caio corrige o nome do arquivo)."""
    base = os.path.splitext(os.path.basename(caminho))[0]
    alvo = normalizar(base)
    # nomes mais longos primeiro, p/ um nome curto não "roubar" o casamento
    candidatos = sorted(PARTICIPANTES, key=lambda p: len(normalizar(p[0])), reverse=True)
    for disp, aba in candidatos:
        if alvo.endswith(normalizar(disp)):
            return aba
    return None


def merge_arquivos(wb, arquivos, primeiro_jogo):
    """Lê cada arquivo preenchido e grava os palpites na aba da mestre do dono.

    O jogo local i (1..N) vira o jogo (primeiro_jogo + i - 1) na mestre — porque
    no molde os jogos são numerados de 1, mas na mestre o mata-mata começa em 73.

    Retorna (gravados, sem_dono):
    - gravados: {aba: nº de jogos gravados}
    - sem_dono: arquivos cujo dono não foi reconhecido (pra avisar o Caio)."""
    gravados = {}
    sem_dono = []
    for caminho in arquivos:
        aba = participante_do_arquivo(caminho)
        if aba is None or aba not in wb.sheetnames:
            sem_dono.append(caminho)
            continue
        wb_arq = openpyxl.load_workbook(caminho)
        local = ler_palpites_arquivo(wb_arq)
        palpites = {primeiro_jogo + (n - 1): gols for n, gols in local.items()}
        escrever_palpites(wb, aba, palpites)
        gravados[aba] = len(palpites)
    return gravados, sem_dono
